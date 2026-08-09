"""XLSX provider — write a real .xlsx workbook from an ExcelWorkbookPayload.

Implements ``adapters/dashboard_render/xlsx.md``. ``openpyxl`` is imported lazily, so
this module imports everywhere; only a real (non-dry-run) write needs the dependency.
``dry_run`` plans the workbook without writing and never needs openpyxl.
"""

from __future__ import annotations

import hashlib
import os
from typing import List, Optional

from ...pipelines.excel_profile import ExcelWorkbookPayload
from .result import FileRecord, RenderResult

# openpyxl chart class name per Excel chart type (others render as a titled cell).
_OPENPYXL_CHART = {
    "columnClustered": ("BarChart", {"type": "col"}),
    "line": ("LineChart", {}),
    "area": ("AreaChart", {}),
    "xyScatter": ("ScatterChart", {}),
    "doughnut": ("DoughnutChart", {}),
}


def _openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def write_xlsx(
    payload: ExcelWorkbookPayload,
    destination: str,
    dry_run: bool = False,
) -> RenderResult:
    """Write a governed ExcelWorkbookPayload to a .xlsx file.

    Creates a data sheet (header row from the payload's measures and dimensions) and a
    dashboard sheet with one chart per panel (supported types become native charts;
    others a titled cell). ``dry_run`` reports the plan without writing (no openpyxl
    needed). Returns a `RenderResult` with the workbook path in its manifest.
    """
    filename = destination if destination.endswith(".xlsx") else os.path.join(
        destination, f"{payload.dashboard_sheet}.xlsx"
    )

    if dry_run:
        return RenderResult(
            adapter_name="dashboard_render",
            provider="xlsx",
            status="planned",
            artifact_uri=filename,
            files=(FileRecord(path=filename, checksum="", bytes=0),),
            dry_run=True,
        )

    if not _openpyxl_available():
        raise RuntimeError(
            "openpyxl is required to write a .xlsx (install the 'dev' extra); "
            "dry_run works without it"
        )

    from openpyxl import Workbook
    from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, Reference, ScatterChart

    chart_classes = {
        "BarChart": BarChart,
        "LineChart": LineChart,
        "AreaChart": AreaChart,
        "ScatterChart": ScatterChart,
        "DoughnutChart": DoughnutChart,
    }

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = payload.data_sheet

    # Header row: dimensions then measures (deterministic, de-duplicated).
    headers: List[str] = []
    for c in payload.charts:
        for d in c.dimensions:
            if d not in headers:
                headers.append(d)
    for m in payload.measures():
        if m not in headers:
            headers.append(m)
    for col, name in enumerate(headers, start=1):
        data_ws.cell(row=1, column=col, value=name)

    dash_ws = wb.create_sheet(title=payload.dashboard_sheet)
    ref = Reference(data_ws, min_col=1, min_row=1, max_col=max(len(headers), 1), max_row=1)
    anchor_row = 1
    for chart in payload.charts:
        dash_ws.cell(row=anchor_row, column=1, value=chart.title)
        spec = _OPENPYXL_CHART.get(chart.chart_type)
        if spec is not None:
            cls_name, attrs = spec
            obj = chart_classes[cls_name]()
            for k, v in attrs.items():
                setattr(obj, k, v)
            obj.title = f"{chart.title} ({chart.measure})"
            obj.add_data(ref, titles_from_data=True)
            dash_ws.add_chart(obj, f"A{anchor_row + 1}")
        else:
            dash_ws.cell(row=anchor_row + 1, column=1, value=f"[{chart.chart_type}] {chart.measure}")
        anchor_row += 15

    os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)
    wb.save(filename)

    with open(filename, "rb") as fh:
        data = fh.read()
    return RenderResult(
        adapter_name="dashboard_render",
        provider="xlsx",
        status="generated",
        artifact_uri=filename,
        files=(FileRecord(path=filename, checksum=hashlib.sha256(data).hexdigest(), bytes=len(data)),),
        dry_run=False,
    )
